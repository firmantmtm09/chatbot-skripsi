import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
csv_path = os.path.join(BASE_DIR, "evaluasi_ragas_final.csv")
excel_path = os.path.join(BASE_DIR, "evaluasi_ragas_final.xlsx")

if not os.path.exists(csv_path):
    print(f"⚠️ Error: File '{csv_path}' tidak ditemukan! Jalankan evaluasi terlebih dahulu.")
    exit()

df = pd.read_csv(csv_path)

wb = openpyxl.Workbook()

ws_dash = wb.active
ws_dash.title = "Ringkasan Evaluasi"
ws_dash.views.sheetView[0].showGridLines = True

ws_dash.column_dimensions['A'].width = 3
ws_dash.column_dimensions['B'].width = 25
ws_dash.column_dimensions['C'].width = 18
ws_dash.column_dimensions['D'].width = 65

ws_dash.cell(row=2, column=2, value="DASHBOARD EVALUASI RAGAS").font = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
ws_dash.cell(row=3, column=2, value="Sistem Question-Answering Chatbot Dukcapil DKI Jakarta").font = Font(name="Segoe UI", size=11, italic=True, color="595959")

metrics = [
    ("Faithfulness", f"='Data Evaluasi Ragas'!E{len(df) + 2}", "Mengukur tingkat keaslian jawaban chatbot berdasarkan dokumen konteks referensi (mendeteksi halusinasi)."),
    ("Answer Relevance", f"='Data Evaluasi Ragas'!F{len(df) + 2}", "Mengukur seberapa relevan, lugas, dan tepat sasaran jawaban chatbot dalam merespons pertanyaan user."),
    ("Context Precision", f"='Data Evaluasi Ragas'!G{len(df) + 2}", "Mengukur akurasi performa Vector Search dalam mengambil potongan dokumen kependudukan yang relevan.")
]

start_r = 6
for title, formula, desc in metrics:
    cell_title = ws_dash.cell(row=start_r, column=2, value=title)
    cell_title.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    cell_title.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    cell_title.alignment = Alignment(horizontal="left", vertical="center")
    
    cell_val = ws_dash.cell(row=start_r, column=3, value=formula)
    cell_val.font = Font(name="Segoe UI", size=14, bold=True, color="1F497D")
    cell_val.alignment = Alignment(horizontal="center", vertical="center")
    cell_val.number_format = "0.00"
    cell_val.border = Border(top=Side(style='thin', color='B0C4DE'), bottom=Side(style='thin', color='B0C4DE'), right=Side(style='thin', color='B0C4DE'), left=Side(style='thin', color='B0C4DE'))
    
    cell_desc = ws_dash.cell(row=start_r, column=4, value=desc)
    cell_desc.font = Font(name="Segoe UI", size=9.5, color="404040")
    cell_desc.alignment = Alignment(vertical="center", wrap_text=True)
    
    ws_dash.row_dimensions[start_r].height = 28
    start_r += 2

ws_data = wb.create_sheet(title="Data Evaluasi Ragas")
ws_data.views.sheetView[0].showGridLines = True

headers = ["No", "Question", "Answer", "Ground Truth", "Faithfulness", "Answer Relevance", "Context Precision"]

for col_idx, header in enumerate(headers, 1):
    cell = ws_data.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

for r_idx, row in df.iterrows():
    actual_row = r_idx + 2
    ws_data.cell(row=actual_row, column=1, value=r_idx + 1).alignment = Alignment(horizontal="center", vertical="top")
    ws_data.cell(row=actual_row, column=2, value=row['question']).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_data.cell(row=actual_row, column=3, value=row['answer']).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_data.cell(row=actual_row, column=4, value=row['ground_truth']).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    for c_idx, metric_col in enumerate(['faithfulness', 'answer_relevance', 'context_precision'], start=5):
        score_cell = ws_data.cell(row=actual_row, column=c_idx, value=float(row[metric_col]))
        score_cell.alignment = Alignment(horizontal="right", vertical="top")
        score_cell.number_format = "0.00"

    for col_idx in range(1, 8):
        ws_data.cell(row=actual_row, column=col_idx).font = Font(name="Segoe UI", size=10)
        ws_data.cell(row=actual_row, column=col_idx).border = thin_border

avg_row_idx = len(df) + 2
ws_data.cell(row=avg_row_idx, column=1, value="Rata-rata Skor Metrik").font = Font(name="Segoe UI", size=11, bold=True)
ws_data.merge_cells(start_row=avg_row_idx, start_column=1, end_row=avg_row_idx, end_column=4)
ws_data.cell(row=avg_row_idx, column=1).alignment = Alignment(horizontal="right", vertical="center")

for col_idx in range(5, 8):
    col_letter = get_column_letter(col_idx)
    cell = ws_data.cell(row=avg_row_idx, column=col_idx, value=f"=AVERAGE({col_letter}2:{col_letter}{avg_row_idx-1})")
    cell.font = Font(name="Segoe UI", size=11, bold=True, color="1F497D")
    cell.number_format = "0.00"
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))

ws_data.column_dimensions['A'].width = 6
ws_data.column_dimensions['B'].width = 35
ws_data.column_dimensions['C'].width = 45
ws_data.column_dimensions['D'].width = 35
ws_data.column_dimensions['E'].width = 15
ws_data.column_dimensions['F'].width = 18
ws_data.column_dimensions['G'].width = 18

ws_data.row_dimensions[1].height = 28
for r in range(2, avg_row_idx + 1):
    ws_data.row_dimensions[r].height = 24

color_scale = ColorScaleRule(start_type='num', start_value=0.0, start_color='FCE4D6',  
                             mid_type='num', mid_value=0.5, mid_color='FFF2CC',    
                             end_type='num', end_value=1.0, end_color='E2EFDA')    
ws_data.conditional_formatting.add(f"E2:G{avg_row_idx-1}", color_scale)

wb.save(excel_path)
print("==================================================")
print("BERHASIL MENGONVERSI DATA KE EXCEL!")
print(f"Laporan rapi disimpan di: {excel_path}")
print("==================================================")